from functools import partial
from openpyxl.cell.read_only import EmptyCell
import pyodbc
import datetime 
import requests
import json
import time
import datetime
import openpyxl

def bill():

    path2 =  "C:\\Users\\Hglaser4\\Documents\\Hg4copy.xlsx"
    wb_obj=openpyxl.Workbook()
    wb_obj = openpyxl.load_workbook(path2)
   
    billing = wb_obj.active

    stdate = input("Enter Start Date YYYY-MM-DD\n")
    endate = input("Enter End Date YYYY-MM-DD\n")
    start =datetime.datetime.strptime(stdate,'%Y-%m-%d').date()
    end= datetime.datetime.strptime(endate,'%Y-%m-%d').date()
    
    max_row_for_a = max((a.row for a in billing['K'] if a.value is not None))

    for y in range(max_row_for_a):
        PostPeriod = billing.cell(row = y+2, column = 11)
        
        if PostPeriod.value is not None:
            PostPeriodDate = PostPeriod.value.date().strftime('%Y-%m-%d')
            start_str = start.strftime('%Y-%m-%d')
            end_str = end.strftime('%Y-%m-%d')

            if start_str <= PostPeriodDate <= end_str:
                print("Nice "+PostPeriodDate)
                #TODO whatever the fuck he wants to do after the date range is found
        
            

   
        
                  
                
       
       
    


   # wb_obj.save(path2)
bill()
